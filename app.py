import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import io
from datetime import datetime

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="Vessel Schedule Workspace", layout="wide")

# --- REGION CONFIGURATIONS ---
REGION_CONFIG = {
    "USA East": ["NY", "Charleston", "Sav", "Norfolk"],
    "USA West": ["LA"],
    "China": ["Shanghai", "Qingdao", "Ningbo"],
    "Europe": ["Hamburg", "Rotterdam", "Antwerp"],
    "Spain": ["Barcelona", "Valencia"],
    "Vietnam": ["Cat Lai"],
    "Indonesia": ["Jakarta"],
    "Mexico": ["Manzanillo"],
    "Malaysia": ["Port Klang", "Penang", "Pasir Gudang"],
    "Bangladesh": ["Chittagong"],
    "Other": ["Colombo"]
}

PORT_REGION_MAP = {
    "NEW YORK": ("USA East", "NY"), "NEWARK": ("USA East", "NY"),
    "CHARLESTON": ("USA East", "Charleston"), "SAVANNAH": ("USA East", "Sav"), "NORFOLK": ("USA East", "Norfolk"),
    "LOS ANGELES": ("USA West", "LA"), "LONG BEACH": ("USA West", "LA"),
    "SHANGHAI": ("China", "Shanghai"), "NINGBO": ("China", "Ningbo"), "QINGDAO": ("China", "Qingdao"),
    "HAMBURG": ("Europe", "Hamburg"), "ROTTERDAM": ("Europe", "Rotterdam"), "ANTWERP": ("Europe", "Antwerp"),
    "BARCELONA": ("Spain", "Barcelona"), "VALENCIA": ("Spain", "Valencia"),
    "CAT LAI": ("Vietnam", "Cat Lai"), "HO CHI MINH": ("Vietnam", "Cat Lai"),
    "JAKARTA": ("Indonesia", "Jakarta"), "MANZANILLO": ("Mexico", "Manzanillo"),
    "PORT KLANG": ("Malaysia", "Port Klang"), "PENANG": ("Malaysia", "Penang"), "PASIR GUDANG": ("Malaysia", "Pasir Gudang"),
    "CHITTAGONG": ("Bangladesh", "Chittagong"), "COLOMBO": ("Other", "Colombo")
}

# --- CUSTOM HTML PARSER (HMM & Fake XLS Files) ---
def parse_html_table(html_text):
    thead = re.search(r'<thead.*?>(.*?)</thead>', html_text, re.IGNORECASE | re.DOTALL)
    if thead:
        headers = re.findall(r'<th.*?>(.*?)</th>', thead.group(1), re.IGNORECASE | re.DOTALL)
        headers = [re.sub(r'<[^>]+>', '', h).strip() for h in headers]
    else:
        first_tr = re.search(r'<tr.*?>(.*?)</tr>', html_text, re.IGNORECASE | re.DOTALL)
        if first_tr:
            headers = re.findall(r'<t[hd].*?>(.*?)</t[hd]>', first_tr.group(1), re.IGNORECASE | re.DOTALL)
            headers = [re.sub(r'<[^>]+>', '', h).strip() for h in headers]
        else:
            headers = []

    tbody = re.search(r'<tbody.*?>(.*?)</tbody>', html_text, re.IGNORECASE | re.DOTALL)
    row_html = tbody.group(1) if tbody else html_text
        
    trs = re.findall(r'<tr.*?>(.*?)</tr>', row_html, re.IGNORECASE | re.DOTALL)
    data = []
    
    start_idx = 0
    if not tbody and len(trs) > 0 and len(headers) > 0:
        first_row_cells = re.findall(r'<t[hd].*?>(.*?)</t[hd]>', trs[0], re.IGNORECASE | re.DOTALL)
        first_row_cleaned = [re.sub(r'<[^>]+>', '', td).strip() for td in first_row_cells]
        if first_row_cleaned == headers:
            start_idx = 1
            
    for tr in trs[start_idx:]:
        tds = re.findall(r'<t[hd].*?>(.*?)</t[hd]>', tr, re.IGNORECASE | re.DOTALL)
        row = [re.sub(r'<[^>]+>', '', td).strip() for td in tds]
        if row and any(row):
            data.append(row)
            
    if data and headers:
        for i in range(len(data)):
            if len(data[i]) < len(headers):
                data[i].extend([''] * (len(headers) - len(data[i])))
            elif len(data[i]) > len(headers):
                data[i] = data[i][:len(headers)]
                
    return pd.DataFrame(data, columns=headers)

# --- ENGINE HELPER FUNCTIONS ---
def normalize_date(val, default_year):
    if pd.isna(val) or not str(val).strip() or str(val).lower() == 'nan': return "-"
    s = str(val).strip()
    
    if 'T' in s: s = s.split('T')[0]
    
    s = re.sub(r'(?i)(CY|SI)\s*Cutoff:', '', s)
    s = re.sub(r'(?i)(Doc|FCL|Port)\s*Cutoff:', '', s)
    s = re.sub(r'(?i)\([a-z]{3}\)', '', s)
    s = re.sub(r'(?i)\b(mon|tue|wed|thu|fri|sat|sun)[a-z]*\b', '', s)
    s = s.replace('()', '').replace('(/)', '')
    s = re.sub(r'\d{1,2}:\d{2}', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    
    if '|' in s: s = s.split('|')[0].strip()
    
    match = re.search(r'(\d{1,2}[-/\s][A-Za-z]{3,9}[-/\s]\d{2,4})|(\d{4}[-/\s]\d{1,2}[-/\s]\d{1,2})', s)
    if match: s_clean = match.group(0)
    else:
        match_no_year = re.search(r'(\d{1,2})[-/\s]([A-Za-z]{3,9})', s)
        if match_no_year and default_year:
            s_clean = f"{match_no_year.group(1)} {match_no_year.group(2)} {default_year}"
        else:
            return "-"

    formats_to_try = [
        '%d %b %Y', '%Y-%m-%d', '%d-%b-%Y', '%d-%B-%Y', 
        '%Y/%m/%d', '%d/%m/%Y', '%d-%b-%y', '%d %B %Y', '%d-%b %Y'
    ]
    for fmt in formats_to_try:
        try: return datetime.strptime(s_clean, fmt).strftime('%d-%b-%y')
        except ValueError: continue
    return "-"

def auto_sort_cutoffs(doc_str, fcl_str):
    """Safety net: Mathematically ensures Doc Cutoff <= FCL Cutoff"""
    if doc_str != "-" and fcl_str != "-":
        try:
            d1 = datetime.strptime(doc_str, '%d-%b-%y')
            d2 = datetime.strptime(fcl_str, '%d-%b-%y')
            if d1 > d2: return fcl_str, doc_str
        except: pass
    return doc_str, fcl_str

def calculate_tts(etd_str, eta_str, default_tts):
    """Calculates true Transit Days based on ETD and ETA dates."""
    if etd_str != "-" and eta_str != "-":
        try:
            etd = datetime.strptime(etd_str, '%d-%b-%y')
            eta = datetime.strptime(eta_str, '%d-%b-%y')
            days = (eta - etd).days
            if days > 0: return str(days)
        except: pass
    return str(default_tts).replace('.0', '')

def clean_vessel_name(name):
    if pd.isna(name) or str(name).lower() == 'nan': return "-"
    return re.sub(r'\[.*?\]', '', str(name).upper()).strip()

def extract_voyage(voy):
    if pd.isna(voy) or str(voy).lower() == 'nan': return "-"
    return str(voy).upper().strip().lstrip('0')

def find_col(df, keywords):
    for c in df.columns:
        if any(k in str(c).lower() for k in keywords): return c
    return None

# --- PARSING ENGINES ---
def parse_file(uploaded_file):
    name = uploaded_file.name.lower()
    rows = []
    dfs = []
    
    try:
        if name.endswith('.csv'):
            dfs = [pd.read_csv(uploaded_file, on_bad_lines='skip')]
        else:
            engine = 'xlrd' if name.endswith('.xls') else 'openpyxl'
            xls = pd.ExcelFile(uploaded_file, engine=engine)
            # Read ALL sheets to prevent missing any vessel mappings
            dfs = [pd.read_excel(xls, sheet_name=s) for s in xls.sheet_names]
            
    except Exception as e:
        error_msg = str(e)
        if "Expected BOF record" in error_msg or "<html" in error_msg.lower() or "<table" in error_msg.lower() or "lxml" in error_msg.lower() or "format cannot be determined" in error_msg.lower():
            try:
                uploaded_file.seek(0)
                html_content = uploaded_file.read().decode('utf-8', errors='ignore')
                dfs = [parse_html_table(html_content)] # Dependency-free HMM unlocker
            except Exception as html_err:
                st.error(f"Failed to read {uploaded_file.name} as an HTML table: {html_err}")
                return []
        else:
            st.error(f"Failed to read {uploaded_file.name}: {e}")
            return []

    for df in dfs:
        cols_str = " ".join([str(c).lower().strip() for c in df.columns])
        sample_blob = df.iloc[:5].to_string().lower()
        year_match = re.search(r'20\d{2}', cols_str + " " + sample_blob)
        file_year = year_match.group(0) if year_match else str(datetime.now().year)

        try:
            # 1. CMA CGM Engine
            if 'vessel name' in cols_str or 'voyage ref' in cols_str:
                v_col = find_col(df, ['vessel name', 'vessel'])
                voy_col = find_col(df, ['voyage ref', 'voyage'])
                dest_col = find_col(df, ['destination', 'arrival location', 'pod'])
                doc_col = find_col(df, ['si cut-off', 'si cutoff', 'doc cut'])
                fcl_col = find_col(df, ['standard booking', 'port cut', 'cy cut', 'vgm cut'])
                etd_col = find_col(df, ['departure date', 'etd'])
                eta_col = find_col(df, ['arrival date', 'eta'])
                tts_col = find_col(df, ['transit time', 'tts'])
                svc_col = find_col(df, ['service'])
                ts_col = find_col(df, ['transhipment', 't/s'])

                for _, r in df.iterrows():
                    v_raw = str(r.get(v_col, '-'))
                    # Skip vertical headers stacked inside the data
                    if str(v_raw).lower() in ['nan', 'vessel name', 'vessel', '-']: continue
                    
                    dest = str(r.get(dest_col, '')).upper()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    
                    service = str(r.get(svc_col, '-')).strip() if svc_col else "-"
                    if service.lower() == 'nan': service = "-"
                    
                    ts_val = str(r.get(ts_col, '')).strip() if ts_col else ""
                    if ts_val in ['1.0', '1', 'Yes', 'Y']:
                        service = f"{service} (T/S)" if service != "-" else "Transshipment"
                    
                    doc_raw = normalize_date(r.get(doc_col, '-'), file_year) if doc_col else "-"
                    fcl_raw = normalize_date(r.get(fcl_col, '-'), file_year) if fcl_col else "-"
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)

                    etd = normalize_date(r.get(etd_col, '-'), file_year) if etd_col else "-"
                    eta = normalize_date(r.get(eta_col, '-'), file_year) if eta_col else "-"
                    
                    tts_val = str(r.get(tts_col, '-')).split('.')[0] if tts_col else "-"
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "CMA CGM", "vessel_name": clean_vessel_name(v_raw), 
                        "voyage_no": extract_voyage(r.get(voy_col, '-')) if voy_col else "-",
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": service
                    })

            # 2. Hapag Lloyd Engine
            elif 'doc cut-off/fcl cut-off/vgm cut-off' in cols_str:
                for _, r in df.iterrows():
                    v_raw = str(r.get('Vessel name', '-'))
                    if str(v_raw).lower() in ['nan', 'vessel name', '-']: continue
                    
                    dest = str(r.get('Arrival location', '')).upper()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    cell = str(r.get('Doc Cut-off/FCL Cut-off/VGM Cut-off', ''))
                    splits = cell.split('/') if '/' in cell else [cell, cell]
                    
                    service = str(r.get('Service name', '-')).strip()
                    if service.lower() == 'nan': service = "-"

                    doc_raw = normalize_date(splits[0], file_year)
                    fcl_raw = normalize_date(splits[1] if len(splits)>1 else splits[0], file_year)
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)

                    etd = normalize_date(r.get('Departure time and date', '-'), file_year)
                    eta = normalize_date(r.get('Arrival time and date', '-'), file_year)
                    tts_val = str(r.get('Transit time', '-')).split('.')[0]
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "Hapag Lloyd", "vessel_name": clean_vessel_name(v_raw), "voyage_no": "-",
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": service
                    })

            # 3. ONE Line Engine
            elif 'unnamed: 10' in cols_str or 'go to one schedule page' in cols_str:
                for i in range(2, len(df)):
                    r = df.iloc[i]
                    vv = str(r.iloc[10] if len(r) > 10 else "").split('\n')[-1].strip()
                    if not vv or vv.lower().startswith('vessel') or str(vv).lower() == 'nan': continue
                    
                    parts = vv.rsplit(' ', 1)
                    dest = str(r.iloc[4] if len(r) > 4 else "").upper()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    
                    service = str(r.iloc[9]).replace('\n', ' ').strip() if len(r) > 9 else "-"
                    if service.lower() == 'nan': service = "-"

                    ts_raw = str(r.iloc[2]).strip() if len(r) > 2 else ""
                    ts_port = re.sub(r'\d{4}-\d{2}-\d{2}.*', '', ts_raw).strip().title()
                    if pd.notna(r.get(2)) and ts_port and ts_port.lower() != "nan":
                        service = f"{service} (T/S {ts_port})" if service != "-" else f"T/S {ts_port}"

                    doc_raw = normalize_date(str(r.iloc[6]).split('\n')[0], file_year) if len(r) > 6 else "-"
                    fcl_raw = normalize_date(str(r.iloc[7]).split('\n')[0], file_year) if len(r) > 7 else "-"
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)

                    etd = normalize_date(str(r.iloc[1]).split('\n')[-1], file_year) if len(r) > 1 else "-"
                    eta = normalize_date(str(r.iloc[4]).split('\n')[-1], file_year) if len(r) > 4 else "-"
                    tts_val = str(r.iloc[5]).split(' ')[0] if len(r) > 5 else "-"
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "ONE Line", "vessel_name": clean_vessel_name(parts[0]), 
                        "voyage_no": extract_voyage(parts[1] if len(parts)>1 else "-"),
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": service
                    })

            # 4. OOCL Engine
            elif 'vessel voyage' in sample_blob or 'unnamed: 8' in cols_str:
                for i in range(len(df) - 1):
                    r_top = df.iloc[i]
                    r_bot = df.iloc[i+1]
                    vv = str(r_top.iloc[8] if len(r_top) > 8 else "").strip()
                    if not vv or vv.lower().startswith('vessel') or str(vv).lower() == 'nan': continue
                    if 'CY Cutoff' not in str(r_top.iloc[4] if len(r_top) > 4 else ""): continue
                    
                    parts = vv.rsplit(' ', 1)
                    dest = str(r_top.iloc[2] if len(r_top) > 2 else "").upper()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    
                    fcl_str = str(r_top.iloc[4] if len(r_top) > 4 else "").replace('CY Cutoff:', '')
                    doc_str = str(r_bot.iloc[4] if len(r_bot) > 4 else "").replace('SI Cutoff:', '')
                    
                    doc_raw = normalize_date(doc_str, file_year)
                    fcl_raw = normalize_date(fcl_str, file_year)
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)
                    
                    service = str(r_top.iloc[7]).strip() if len(r_top) > 7 else "-"
                    if service.lower() == 'nan': service = "-"

                    ts_raw = str(r_top.iloc[1]).strip().title() if len(r_top) > 1 else ""
                    if pd.notna(r_top.get(1)) and ts_raw and ts_raw.lower() != "nan":
                        service = f"{service} (T/S {ts_raw})" if service != "-" else f"T/S {ts_raw}"

                    etd = normalize_date(r_bot.iloc[5] if len(r_bot) > 5 else "-", file_year)
                    eta = normalize_date(r_bot.iloc[2] if len(r_bot) > 2 else "-", file_year)
                    tts_val = str(r_top.iloc[3]).split(' ')[0] if len(r_top) > 3 else "-"
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "OOCL", "vessel_name": clean_vessel_name(parts[0]), 
                        "voyage_no": extract_voyage(parts[1] if len(parts)>1 else "-"),
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": service
                    })

            # 5. Maersk Engine
            elif 'deadline cy' in cols_str:
                v_col = find_col(df, ['vessel'])
                voy_col = find_col(df, ['voyage number', 'voyage'])
                for _, r in df.iterrows():
                    v_raw = str(r.get(v_col, '-'))
                    if str(v_raw).lower() in ['nan', 'vessel', '-']: continue

                    dest = str(r.get('Arrival', '')).upper().split('-')[0].strip()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    
                    doc_raw = normalize_date(r.get('Deadline SI - NON AMS', '-'), file_year)
                    fcl_raw = normalize_date(r.get('Deadline CY', '-'), file_year)
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)

                    etd = normalize_date(r.get('Departure Date', '-'), file_year)
                    eta = normalize_date(r.get('Arrival Date', '-'), file_year)
                    tts_val = str(r.get('Transit Time', '-')).split(' ')[0]
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "Maersk", "vessel_name": clean_vessel_name(v_raw), 
                        "voyage_no": extract_voyage(r.get(voy_col, '-')),
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": "-"
                    })

            # 6. HMM Engine
            elif 'total transittime(days)' in cols_str or 's/i cut-off' in cols_str:
                for _, r in df.iterrows():
                    vessel_raw = str(r.get('Vessel', '-'))
                    if str(vessel_raw).lower() in ['nan', 'vessel', '-']: continue

                    dest = str(r.get('DischargingPort', '')).upper()
                    reg, pod = next((v for k, v in PORT_REGION_MAP.items() if k in dest), ("USA East", "NY"))
                    
                    service = str(r.get('Route', '-')).strip()
                    if service.lower() == 'nan': service = "-"

                    ts_raw = str(r.get('Next Port(T/S)', '')).strip().title()
                    if pd.notna(r.get('Next Port(T/S)')) and ts_raw and ts_raw.lower() != "nan":
                        service = f"{service} (T/S {ts_raw})" if service != "-" else f"T/S {ts_raw}"

                    first_vessel = vessel_raw.split('\t')[0]
                    first_vessel = re.sub(r'\[.*?\]', '', first_vessel).strip()
                    v_parts = first_vessel.rsplit(' ', 1)
                    v_name = clean_vessel_name(v_parts[0])
                    voy_no = extract_voyage(v_parts[1] if len(v_parts)>1 else "-")

                    etd_raw = str(r.get('Loading Port', ''))
                    etd_match = re.search(r'ETD\s*:\s*(\d{4}-\d{2}-\d{2})', etd_raw)
                    etd = normalize_date(etd_match.group(1) if etd_match else etd_raw, file_year)

                    eta_raw = str(r.get('DischargingPort', ''))
                    eta_match = re.search(r'ETB\s*:\s*(\d{4}-\d{2}-\d{2})', eta_raw)
                    eta = normalize_date(eta_match.group(1) if eta_match else eta_raw, file_year)

                    doc_raw = normalize_date(str(r.get('S/I Cut-off', '-')).split('T')[0], file_year)
                    fcl_raw = normalize_date(str(r.get('Cargo Cut-off', '-')).split('T')[0], file_year)
                    doc_clean, fcl_clean = auto_sort_cutoffs(doc_raw, fcl_raw)

                    tts_val = str(r.get('Total TransitTime(Days)', '-')).split('.')[0]
                    tts_final = calculate_tts(etd, eta, tts_val)

                    rows.append({
                        "region": reg, "carrier": "HMM", "vessel_name": v_name, "voyage_no": voy_no,
                        "doc_cutoff": doc_clean, "fcl_cutoff": fcl_clean,
                        "etd": etd, "etas": {pod: eta},
                        "tts": {pod: tts_final}, "service": service
                    })
        except Exception as e:
            st.warning(f"Error parsing layout block in {name}: {e}")
            
    return rows

# --- FRONTEND UI ---
st.title("🚢 Master Vessel Schedule Aggregator")
st.markdown("Drag and drop raw carrier exports. The engine will automatically identify the carrier, fix the dates, calculate exact transit days, and build the 9-tab corporate Excel file.")

uploaded_files = st.file_uploader("Upload Carrier Exports (.xlsx, .xls, .csv)", accept_multiple_files=True)

if uploaded_files:
    database = {}
    with st.spinner('Parsing matrices and aligning timelines...'):
        for f in uploaded_files:
            extracted = parse_file(f)
            for r in extracted:
                v_key = re.sub(r'[^A-Z0-9]', '', r['vessel_name'])
                voy_key = re.sub(r'[^A-Z0-9]', '', r['voyage_no'])
                comp_key = f"{r['region']}|||{r['carrier']}|||{v_key}|||{voy_key}"
                
                if comp_key not in database:
                    database[comp_key] = r
                else:
                    database[comp_key]['etas'].update(r['etas'])
                    database[comp_key]['tts'].update(r['tts'])
                    if database[comp_key]['doc_cutoff'] == "-": database[comp_key]['doc_cutoff'] = r['doc_cutoff']
                    if database[comp_key]['fcl_cutoff'] == "-": database[comp_key]['fcl_cutoff'] = r['fcl_cutoff']
                    if database[comp_key].get('service', '-') == "-" and r.get('service', '-') != "-": 
                        database[comp_key]['service'] = r['service']

    all_items = list(database.values())
    st.success(f"✅ Successfully compiled {len(all_items)} unique vessel rotations!")
    
    # --- BUILD EXCEL FILE IN MEMORY ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    fill_header = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    fill_meta = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
    fill_sop = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for reg_name, ports in REGION_CONFIG.items():
        ws = wb.create_sheet(title=reg_name)
        ws.append(["JUNE 2026 VESSEL SCHEDULE"])
        ws.append(["STANDARD OPERATING PARAMETERS"])
        ws.append(["· Gate Open: 5 Days Prior to Vessel ETD"])
        ws.append(["· Gate Close: 2 Days Prior to Vessel ETD"])
        ws.append(["· Documentation Cut-Off: 48 Hours Prior to Vessel ETD"])
        ws.append([])
        ws.append([f"{reg_name.upper()} MATRIX OVERVIEW"])
        ws.append(["POL: Mundra / Nhava Sheva"])
        ws.append([f"POD Target Gateways: {', '.join(ports)}"])
        ws.append([])
        
        headers = ["Carrier", "Vessel Name", "Voyage No.", "Doc Cut-Off", "FCL Cut-Off", "ETD Local"]
        for p in ports: headers.extend([f"ETA {p}", f"{p} TT"])
        headers.append("Service")
        ws.append(headers)
        
        reg_rows = [r for r in all_items if r['region'] == reg_name]
        def sort_key(x):
            try: dt = datetime.strptime(x['etd'], '%d-%b-%y')
            except: dt = datetime(9999, 12, 31)
            return (x['carrier'].lower(), dt)
            
        reg_rows.sort(key=sort_key)

        for r in reg_rows:
            cells = [r['carrier'], r['vessel_name'], r['voyage_no'], r['doc_cutoff'], r['fcl_cutoff'], r['etd']]
            for p in ports: cells.extend([r['etas'].get(p, "-"), r['tts'].get(p, "-")])
            cells.append(r.get('service', '-'))
            ws.append(cells)

        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)), 1):
            if r_idx in [1, 2, 7, 8, 9]:
                for cell in row: 
                    cell.fill = fill_meta
                    cell.font = Font(bold=True)
                    cell.alignment = align_left
            elif r_idx in [3, 4, 5]:
                for cell in row: 
                    cell.fill = fill_sop
                    cell.alignment = align_left
            elif r_idx == 11:
                for cell in row: 
                    cell.fill = fill_header
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = align_center
            elif r_idx > 11:
                for cell in row: 
                    cell.border = border
                    cell.alignment = align_center
                
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_len: max_len = len(line)
            adjusted_width = min(max_len + 4, 45)
            adjusted_width = max(adjusted_width, 14)
            ws.column_dimensions[col_letter].width = adjusted_width

        if ws.max_row >= 12:
            start_merge_row = 12
            current_carrier = ws.cell(row=12, column=1).value
            
            for r in range(13, ws.max_row + 2):
                cell_val = ws.cell(row=r, column=1).value if r <= ws.max_row else None
                if cell_val != current_carrier:
                    if r - 1 > start_merge_row:
                        ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=r-1, end_column=1)
                    start_merge_row = r
                    current_carrier = cell_val

    out = io.BytesIO()
    wb.save(out)
    
    st.markdown("---")
    st.download_button(
        label="📥 Download Compiled Master Excel Sheet",
        data=out.getvalue(),
        file_name=f"{datetime.now().strftime('%B')} Vessel Schedule Master.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
